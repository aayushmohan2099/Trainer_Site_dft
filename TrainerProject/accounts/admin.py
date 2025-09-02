from django.contrib import admin

from .models import Beneficiary, MasterTrainer

from django import forms
from django.shortcuts import render, redirect
from django.urls import path
import openpyxl

from django.http import HttpResponse

from django.utils.html import format_html
from django.urls import reverse

from django.contrib.auth.models import User

import random, string
from django.contrib.auth.hashers import make_password

# Excel -> DB Upload Form
class ExcelUploadForm(forms.Form):
    file = forms.FileField()

# Generating Random Usernames and Passwords for new users

def generate_username(prefix, num):
    return f"{prefix}{num}{random.randint(1000,9999)}"

def generate_password(length=8):
    chars = string.ascii_letters + string.digits + "!@#$%^&*"
    return ''.join(random.choice(chars) for _ in range(length))

# Custom Admin Portal for Import/Export of Bulk Data

# Beneficiary Data
class BeneficiaryAdmin(admin.ModelAdmin):
    list_display = ('member_name', 'shg_name', 'village', 'marital_status')
    change_list_template = "admin/beneficiary_changelist.html"

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('import-excel/', self.admin_site.admin_view(self.import_excel), name="beneficiary_import_excel"),
            path('export-excel/', self.admin_site.admin_view(self.export_excel), name="beneficiary_export_excel"),
            path('download-format/', self.admin_site.admin_view(self.download_format), name="beneficiary_download_format"),
        ]
        return custom_urls + urls

    def import_excel(self, request):
        if request.method == "POST":
            form = ExcelUploadForm(request.POST, request.FILES)
            if form.is_valid():
                wb = openpyxl.load_workbook(form.cleaned_data['file'])
                sheet = wb.active
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    # Random credentials for new users
                    username = generate_username("benef", row[0])
                    password = generate_password()

                    # If Not exist, Create 1 or fetch if already exists
                    user, created = User.objects.get_or_create(
                        username=username,
                        defaults={
                            "email": f"{username}@example.com",
                            "password": make_password(password),
                            "first_name": row[9].split(" ")[0] if row[9] else "",
                            "last_name": row[9].split(" ")[-1] if row[9] else "",
                        }
                    )

                    Beneficiary.objects.create(
                        user=user,
                        state=row[1], district=row[2], block=row[3],
                        gram_panchayat=row[4], village=row[5],
                        shg_code=row[6], shg_name=row[7],
                        member_code=row[8], member_name=row[9],
                        marital_status=row[10], disability_status=row[11],
                        bank_account_no=row[12], ifsc_code=row[13]
                    )

                self.message_user(request, "Beneficiaries imported successfully!")
                return redirect("..")
        else:
            form = ExcelUploadForm()
        return render(request, "admin/upload_excel.html", {"form": form, "title": "Import Beneficiaries"})

    def export_excel(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Beneficiaries"
        ws.append(["UserID", "State", "District", "Block", "GP", "Village",
                   "SHG Code", "SHG Name", "Member Code", "Member Name",
                   "Marital Status", "Disability", "Bank Account", "IFSC Code"])
        for b in Beneficiary.objects.all():
            ws.append([
                b.user.id, b.state, b.district, b.block, b.gram_panchayat, b.village,
                b.shg_code, b.shg_name, b.member_code, b.member_name,
                b.marital_status, b.disability_status, b.bank_account_no, b.ifsc_code
            ])
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = 'attachment; filename="beneficiaries.xlsx"'
        wb.save(response)
        return response

# Download format for data input
    def download_format(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Beneficiary Format"
        ws.append(["UserID", "State", "District", "Block", "GP", "Village",
                "SHG Code", "SHG Name", "Member Code", "Member Name",
                "Marital Status", "Disability", "Bank Account", "IFSC Code"])
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = 'attachment; filename="beneficiary_format.xlsx"'
        wb.save(response)
        return response

# Master Trainer Data
class MasterTrainerAdmin(admin.ModelAdmin):
    list_display = ('full_name', 'qualification', 'availability')
    change_list_template = "admin/trainer_changelist.html"

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('import-excel/', self.admin_site.admin_view(self.import_excel), name="trainer_import_excel"),
            path('export-excel/', self.admin_site.admin_view(self.export_excel), name="trainer_export_excel"),
            path('download-format/', self.admin_site.admin_view(self.download_format), name="trainer_download_format"),
        ]
        return custom_urls + urls

    def import_excel(self, request):
        if request.method == "POST":
            form = ExcelUploadForm(request.POST, request.FILES)
            if form.is_valid():
                wb = openpyxl.load_workbook(form.cleaned_data['file'])
                sheet = wb.active
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    # Random credentials for new users
                    username = generate_username("trainer", row[0])
                    password = generate_password()

                    # If Not exist, Create 1 or fetch if already exists                    
                    user, created = User.objects.get_or_create(
                        username=username,
                        defaults={
                            "email": f"{username}@example.com",
                            "password": make_password(password),
                            "first_name": row[9].split(" ")[0] if row[9] else "",
                            "last_name": row[9].split(" ")[-1] if row[9] else "",
                        }
                    )

                    MasterTrainer.objects.create(
                        user=user,
                        full_name=row[1],
                        qualification=row[2],
                        expertise=row[3],
                        training_history=row[4],
                        availability=row[5]
                    )
                self.message_user(request, "Trainers imported successfully!")
                return redirect("..")
        else:
            form = ExcelUploadForm()
        return render(request, "admin/upload_excel.html", {"form": form, "title": "Import Trainers"})

    def export_excel(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Trainers"
        ws.append(["UserID", "Full Name", "Qualification", "Expertise", "Training History", "Availability"])
        for t in MasterTrainer.objects.all():
            ws.append([
                t.user.id, t.full_name, t.qualification, t.expertise,
                t.training_history, t.availability
            ])
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = 'attachment; filename="trainers.xlsx"'
        wb.save(response)
        return response

# For bulk data input format download
    def download_format(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Trainer Format"
        ws.append(["UserID", "Full Name", "Qualification", "Expertise", "Training History", "Availability"])
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = 'attachment; filename="trainer_format.xlsx"'
        wb.save(response)
        return response
    

admin.site.register(Beneficiary, BeneficiaryAdmin)
admin.site.register(MasterTrainer, MasterTrainerAdmin)